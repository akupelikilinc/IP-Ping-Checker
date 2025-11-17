import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
import subprocess
import platform
import socket
from datetime import datetime
import os
import time

class IPPingChecker:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.wb = None
        self.ws = None
        self.results = {"yanıt_var": 0, "yanıt_yok": 0}
        
    def create_sample_file(self):
        """Örnek Excel dosyası oluşturur"""
        if not os.path.exists(self.excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "IP Taraması"
            
            # Başlıkları ekle
            ws['A1'] = "IP Adresi"
            ws['B1'] = "Ping"
            ws['C1'] = "Yanıt Süresi"
            ws['D1'] = "Cihaz Adı"
            
            # Başlık formatı
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col in ['A1', 'B1', 'C1', 'D1']:
                ws[col].fill = header_fill
                ws[col].font = header_font
                ws[col].alignment = Alignment(horizontal="center", vertical="center")
            
            # Örnek IP adresleri ekle
            sample_ips = ["8.8.8.8", "1.1.1.1", "192.168.1.1", "192.168.1.2", "10.0.0.1"]
            for idx, ip in enumerate(sample_ips, start=2):
                ws[f'A{idx}'] = ip
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25
            
            wb.save(self.excel_file)
            print(f"✓ Örnek Excel dosyası oluşturuldu: {self.excel_file}")
            return True
        return False
    
    def load_workbook(self):
        """Excel dosyasını yükle"""
        try:
            self.wb = openpyxl.load_workbook(self.excel_file)
            self.ws = self.wb.active
            return True
        except Exception as e:
            print(f"✗ Hata: Excel dosyası yüklenemedi - {e}")
            return False
    
    def ping_host(self, ip_address):
        """Verilen IP adresine ping at ve sonuç döndür"""
        try:
            # Windows için
            if platform.system().lower() == "windows":
                result = subprocess.run(
                    ["ping", "-n", "1", "-w", "1000", ip_address],
                    capture_output=True,
                    text=True,
                    timeout=5
                )
            else:
                # Linux/Mac için
                result = subprocess.run(
                    ["ping", "-c", "1", "-W", "1", ip_address],
                    capture_output=True,
                    text=True,
                    timeout=5
                )
            
            if result.returncode == 0:
                # Yanıt süresi çıkar
                response_time = self._extract_response_time(result.stdout)
                return True, response_time
            else:
                return False, None
                
        except Exception as e:
            print(f"✗ Ping hatası ({ip_address}): {e}")
            return False, None
    
    def _extract_response_time(self, ping_output):
        """Ping çıktısından yanıt süresini çıkar"""
        try:
            # Windows çıktısı örneği: "time=12ms" veya "time<1ms"
            if "time=" in ping_output:
                for line in ping_output.split('\n'):
                    if "time=" in line:
                        start = line.find("time=") + 5
                        end = line.find("ms", start) + 2
                        return line[start:end].strip()
            # Linux/Mac çıktısı
            elif "time=" in ping_output:
                import re
                match = re.search(r'time=(\d+\.?\d*)\s*ms', ping_output)
                if match:
                    return f"{match.group(1)}ms"
        except:
            pass
        return None
    
    def get_device_name(self, ip_address):
        """IP adresinin cihaz adını bulmaya çalışır"""
        try:
            hostname = socket.gethostbyaddr(ip_address)[0]
            return hostname
        except:
            return "-"
    
    def run_ping_check(self):
        """IP adresleri için ping kontrolü çalıştır"""
        if not self.load_workbook():
            return False
        
        self.results = {"yanıt_var": 0, "yanıt_yok": 0}
        
        # Başlık satırından başla (satır 2 ve sonrası)
        row_num = 2
        
        while row_num <= self.ws.max_row:
            ip_cell = self.ws[f'A{row_num}']
            
            # IP adresi boşsa, son satıra gelmiş demektir
            if not ip_cell.value:
                break
            
            ip_address = str(ip_cell.value).strip()
            
            # Geçersiz IP'leri atla
            if not self._is_valid_ip(ip_address):
                print(f"⚠ Geçersiz IP: {ip_address}")
                row_num += 1
                continue
            
            print(f"🔄 Ping atılıyor: {ip_address}...", end=" ")
            
            # Ping at
            is_alive, response_time = self.ping_host(ip_address)
            
            # Sonuçları hücrelere yaz
            ping_cell = self.ws[f'B{row_num}']
            time_cell = self.ws[f'C{row_num}']
            device_cell = self.ws[f'D{row_num}']
            
            if is_alive:
                ping_cell.value = "Yanıt Var"
                ping_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                ping_cell.font = Font(bold=True, color="000000")
                if response_time:
                    time_cell.value = response_time
                self.results["yanıt_var"] += 1
                print(f"✓ Yanıt Var {response_time if response_time else ''}")
            else:
                ping_cell.value = "Yanıt Yok"
                ping_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ping_cell.font = Font(bold=True, color="FFFFFF")
                time_cell.value = "-"
                self.results["yanıt_yok"] += 1
                print(f"✗ Yanıt Yok")
            
            # Cihaz adını bul
            device_cell.value = self.get_device_name(ip_address)
            
            # Merkez hizala
            for cell in [ping_cell, time_cell, device_cell]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1
        
        # Dosyayı kaydet
        self.wb.save(self.excel_file)
        print(f"\n✓ Sonuçlar kaydedildi")
        return True
    
    def _is_valid_ip(self, ip_string):
        """IP adresinin geçerli olup olmadığını kontrol et"""
        parts = ip_string.split('.')
        if len(parts) != 4:
            return False
        try:
            return all(0 <= int(part) <= 255 for part in parts)
        except:
            return False
    
    def create_charts(self):
        """Sonuçlar için grafikler oluştur"""
        if not self.load_workbook():
            return False
        
        # Grafik verilerini hazırla
        chart_sheet = self.wb.create_sheet("Grafikler")
        
        # Özet veri
        chart_sheet['A1'] = "Durum"
        chart_sheet['B1'] = "Sayı"
        chart_sheet['A2'] = "Yanıt Var"
        chart_sheet['B2'] = self.results["yanıt_var"]
        chart_sheet['A3'] = "Yanıt Yok"
        chart_sheet['B3'] = self.results["yanıt_yok"]
        
        # Pasta Grafiği
        pie = PieChart()
        pie.title = "IP Ping Sonuçları (Pasta Grafiği)"
        pie.style = 10
        labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=3)
        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        chart_sheet.add_chart(pie, "A5")
        
        # Bar Grafiği
        bar = BarChart()
        bar.title = "IP Ping Sonuçları (Bar Grafiği)"
        bar.style = 10
        bar.x_axis.title = "Durum"
        bar.y_axis.title = "Sayı"
        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=3)
        cat = Reference(chart_sheet, min_col=1, min_row=2, max_row=3)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cat)
        chart_sheet.add_chart(bar, "A20")
        
        self.wb.save(self.excel_file)
        print("✓ Grafikler oluşturuldu")
        return True
    
    def show_summary(self):
        """Özet bilgi göster"""
        total = self.results["yanıt_var"] + self.results["yanıt_yok"]
        if total == 0:
            print("⚠ Kontrol edilecek IP bulunamadı!")
            return
        
        success_rate = (self.results["yanıt_var"] / total) * 100
        
        print("\n" + "="*50)
        print("📊 PING KONTROLÜ ÖZETI")
        print("="*50)
        print(f"Zaman: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
        print(f"Toplam IP: {total}")
        print(f"✓ Yanıt Var: {self.results['yanıt_var']} (%{success_rate:.1f})")
        print(f"✗ Yanıt Yok: {self.results['yanıt_yok']} (%{100-success_rate:.1f})")
        print("="*50 + "\n")


def main():
    # Excel dosyası yolu - environment variable'dan al veya varsayılan kullan
    import os
    workdir = os.getenv('WORKDIR', '/app/data')
    excel_filename = os.getenv('EXCEL_FILENAME', 'IP_Ping_Sonuclari.xlsx')
    excel_file = os.path.join(workdir, excel_filename)
    
    # Kontrol sınıfı oluştur
    checker = IPPingChecker(excel_file)
    
    # Dosya yoksa örnek oluştur
    checker.create_sample_file()
    
    # Ping kontrolü yap
    if checker.run_ping_check():
        # Grafikler oluştur
        checker.create_charts()
        # Özet göster
        checker.show_summary()
    else:
        print("✗ Ping kontrolü başarısız oldu!")


if __name__ == "__main__":
    main()
